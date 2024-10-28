import requests
import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import azure.functions as func

urls = [
    "https://cir-reports.cir-safety.org/FetchCIRReports",
    "https://cir-reports.cir-safety.org/FetchCIRReports/?&pagingcookie=%26lt%3bcookie%20page%3d%26quot%3b1%26quot%3b%26gt%3b%26lt%3bpcpc_name%20last%3d%26quot%3bPEG-50%20Stearate%26quot%3b%20first%3d%26quot%3b1%2c10-Decanediol%26quot%3b%20%2f%26gt%3b%26lt%3bpcpc_ingredientidname%20last%3d%26quot%3bPEG-50%20Stearate%26quot%3b%20first%3d%26quot%3b1%2c10-Decanediol%26quot%3b%20%2f%26gt%3b%26lt%3bpcpc_cirrelatedingredientsid%20last%3d%26quot%3b%7bC223037E-F278-416D-A287-2007B9671D0C%7d%26quot%3b%20first%3d%26quot%3b%7b940AF697-52B5-4A3A-90A6-B9DB30EF4A7E%7d%26quot%3b%20%2f%26gt%3b%26lt%3b%2fcookie%26gt%3b&page=2",
]


def fetch_cir_reports(url):
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Errore durante il recupero dei dati: {response.status_code}")
        return None

def extract_data_from_json(data):
    records = []
    if data:
        results = data.get("results", [])
        for result in results:
            ingredient_name = result.get("pcpc_ingredientname", "")
            inci_name = result.get("pcpc_ciringredientname", "")
            id_link = result.get("pcpc_ingredientid", "")
            link = f"https://cir-reports.cir-safety.org/cir-ingredient-status-report/?id={id_link}"
            records.append({
                "Ingredient_Name": ingredient_name,
                "INCI_Name": inci_name,
                "Link": link
            })
    else:
        print("Nessun dato disponibile")
    return records

def update_excel_file(new_records, file_path):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)
        wb.save(file_path)
        new_df = pd.DataFrame(new_records)
        new_df.drop_duplicates(inplace=True)
        new_df.to_excel(file_path, index=False)
    else:
        new_df = pd.DataFrame(new_records)
        new_df.drop_duplicates(inplace=True)
        new_df.to_excel(file_path, index=False)

def miao():
    file_path = "/tmp/cir_reports.xlsx"
    new_records = fetch_and_extract()
    if new_records:
        update_excel_file(new_records, file_path)
        remove_duplicates_excel(file_path)
        remove_first_row_if_needed(file_path)

def main(req: func.HttpRequest) -> func.HttpResponse:
    miao()
    return func.HttpResponse("Scraping e aggiornamento del file completati.", status_code=200)
